import io
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modulos.equipos import cargar_equipos
from modulos.inventario import cargar_inventario
from modulos.ordenes import cargar_ordenes

# --- Paleta de colores institucional para Dashboards de Planta ---
VERDE    = "#2ecc71"
AZUL     = "#3498db"
ROJO     = "#e74c3c"
NARANJA  = "#e67e22"
MORADO   = "#9b59b6"
GRIS     = "#95a5a6"

PALETA_BARRAS = [AZUL, VERDE, NARANJA, MORADO, ROJO]


# =============================================================================
# CAPA DE LÓGICA / KPIs (Procesamiento de DataFrames Relacionales)
# =============================================================================

def filtrar_por_periodo(df_ot: pd.DataFrame, fecha_inicio: date, fecha_fin: date) -> pd.DataFrame:
    """Filtra el DataFrame de OTs por el rango de fechas seleccionado en la UI."""
    if df_ot.empty:
        return df_ot

    df = df_ot.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    mascara = (df["Fecha"].dt.date >= fecha_inicio) & (df["Fecha"].dt.date <= fecha_fin)
    return df[mascara].reset_index(drop=True)


def calcular_disponibilidad(df_eq: pd.DataFrame, df_ot: pd.DataFrame) -> dict:
    """Calcula el índice de disponibilidad de la flota en tiempo real (OTs en Ejecución)."""
    total = len(df_eq)
    if total == 0:
        return {"total": 0, "disponibles": 0, "en_mantenimiento": 0, "pct": 0.0}

    equipos_en_mant = set()
    if not df_ot.empty:
        equipos_en_mant = set(
            df_ot[df_ot["Estado"] == "En Ejecución"]["Codigo_Equipo"].str.upper().unique()
        )

    en_mantenimiento = len(equipos_en_mant)
    disponibles = max(total - en_mantenimiento, 0)
    pct = round(disponibles / total * 100, 1)

    return {
        "total": total,
        "disponibles": disponibles,
        "en_mantenimiento": en_mantenimiento,
        "pct": pct,
    }


def consumo_repuestos(df_ot: pd.DataFrame, df_inv: pd.DataFrame, top_n: int = 5) -> pd.DataFrame:
    """Genera el ranking de los repuestos más consumidos en OTs con estado Finalizado."""
    ots_fin = df_ot[df_ot["Estado"] == "Finalizado"] if not df_ot.empty else pd.DataFrame()

    if ots_fin.empty:
        return pd.DataFrame(columns=["Codigo_Repuesto", "Nombre", "Unidad", "Total_Consumido"])

    consumo = (
        ots_fin.groupby("Codigo_Repuesto")["Cantidad"]
        .sum()
        .reset_index()
        .rename(columns={"Cantidad": "Total_Consumido"})
    )

    if not df_inv.empty:
        ref = df_inv[["Codigo", "Nombre", "Unidad"]].copy()
        consumo = consumo.merge(ref, left_on="Codigo_Repuesto", right_on="Codigo", how="left")
        consumo["Nombre"] = consumo["Nombre"].fillna(consumo["Codigo_Repuesto"])
        consumo["Unidad"] = consumo["Unidad"].fillna("")
        consumo = consumo[["Codigo_Repuesto", "Nombre", "Unidad", "Total_Consumido"]]
    else:
        consumo["Nombre"] = consumo["Codigo_Repuesto"]
        consumo["Unidad"] = ""

    return (
        consumo.sort_values("Total_Consumido", ascending=False)
        .head(top_n)
        .reset_index(drop=True)
    )


def ratio_mantenimientos(df_ot: pd.DataFrame) -> dict:
    """Calcula el ratio porcentual de OTs Preventivas vs Correctivas en el período."""
    if df_ot.empty:
        return {"preventivo": 0, "correctivo": 0, "total": 0, "pct_prev": 0.0, "pct_corr": 0.0}

    preventivo = int((df_ot["Tipo_Mantenimiento"] == "Preventivo").sum())
    correctivo = int((df_ot["Tipo_Mantenimiento"] == "Correctivo").sum())
    total = preventivo + correctivo

    return {
        "preventivo": preventivo,
        "correctivo": correctivo,
        "total": total,
        "pct_prev": round(preventivo / total * 100, 1) if total > 0 else 0.0,
        "pct_corr": round(correctivo / total * 100, 1) if total > 0 else 0.0,
    }


# =============================================================================
# APORTE INTEGRANTE 1: CAPA DE VISUALIZACIÓN DINÁMICA (Plotly Engine)
# =============================================================================

def grafico_dona_disponibilidad(metricas: dict) -> go.Figure:
    """Diseño del gráfico de dona con la disponibilidad actual de la flota pesada."""
    labels = ["Disponibles", "En Mantenimiento"]
    values = [metricas["disponibles"], metricas["en_mantenimiento"]]
    colors = [VERDE, AZUL]

    if sum(values) == 0:
        values = [1, 0]

    fig = go.Figure(data=[go.Pie(
        labels=labels,
        values=values,
        hole=0.62,
        marker=dict(colors=colors, line=dict(color="#ffffff", width=2)),
        textinfo="label+value",
        hovertemplate="%{label}: %{value} equipo(s)<extra></extra>",
    )])

    fig.update_layout(
        title=dict(text="Disponibilidad de Flota", font=dict(size=15)),
        annotations=[dict(
            text=f"<b>{metricas['pct']}%</b>",
            font=dict(size=26, color=VERDE),
            showarrow=False,
        )],
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=-0.2),
        margin=dict(t=50, b=50, l=20, r=20),
        height=320,
    )
    return fig


def grafico_dona_mantenimientos(ratio: dict) -> go.Figure:
    """Diseño del gráfico de dona con la distribución de Mantenimientos MP vs MC."""
    labels = ["Preventivo", "Correctivo"]
    values = [ratio["preventivo"], ratio["correctivo"]]
    colors = [VERDE, ROJO]

    if sum(values) == 0:
        values = [1, 0]

    fig = go.Figure(data=[go.Pie(
        labels=labels,
        values=values,
        hole=0.62,
        marker=dict(colors=colors, line=dict(color="#ffffff", width=2)),
        textinfo="label+value",
        hovertemplate="%{label}: %{value} OT(s) (%{percent})<extra></extra>",
    )])

    fig.update_layout(
        title=dict(text="Tipo de Mantenimiento", font=dict(size=15)),
        annotations=[dict(
            text=f"<b>{ratio['pct_prev']}%</b><br>prev.",
            font=dict(size=20, color=VERDE),
            showarrow=False,
        )],
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=-0.2),
        margin=dict(t=50, b=50, l=20, r=20),
        height=320,
    )
    return fig


def grafico_barras_repuestos(df_consumo: pd.DataFrame) -> go.Figure:
    """Diseño del gráfico de barras horizontales de consumo logístico."""
    if df_consumo.empty:
        fig = go.Figure()
        fig.update_layout(
            title="Repuestos más consumidos (OTs Finalizadas)",
            annotations=[dict(
                text="Sin datos de OTs finalizadas en el período",
                showarrow=False,
                font=dict(size=14, color=GRIS),
            )],
            height=300,
        )
        return fig

    etiquetas = df_consumo.apply(
        lambda r: f"{r['Nombre']} ({r['Codigo_Repuesto']})", axis=1
    ).tolist()
    valores = df_consumo["Total_Consumido"].tolist()
    colores = [PALETA_BARRAS[i % len(PALETA_BARRAS)] for i in range(len(etiquetas))]

    fig = go.Figure(data=[go.Bar(
        x=valores,
        y=etiquetas,
        orientation="h",
        marker=dict(color=colores, line=dict(color="#ffffff", width=1)),
        text=[f"{v:.0f}" for v in valores],
        textposition="outside",
        hovertemplate="%{y}: %{x:.0f} unidades<extra></extra>",
    )])

    fig.update_layout(
        title=dict(text="Repuestos más consumidos (OTs Finalizadas)", font=dict(size=15)),
        xaxis_title="Cantidad total consumida",
        yaxis=dict(autorange="reversed"),
        margin=dict(t=50, b=40, l=20, r=60),
        height=max(280, len(etiquetas) * 55),
        plot_bgcolor="#fafafa",
    )
    return fig


# =============================================================================
# APORTE INTEGRANTE 2: ENGINE DE EXPORTACIÓN Y FORMATO CORPORATIVO (openpyxl)
# =============================================================================

def _estilos_excel() -> dict:
    """Devuelve los estilos visuales estandarizados para reportes de planta."""
    borde_celda = Border(
        left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"),
    )
    return {
        "fill_titulo":   PatternFill("solid", fgColor="1F4E79"),
        "fill_seccion":  PatternFill("solid", fgColor="2E75B6"),
        "fill_col_hdr":  PatternFill("solid", fgColor="BDD7EE"),
        "fill_alt":      PatternFill("solid", fgColor="DEEAF1"),
        "fill_fin":      PatternFill("solid", fgColor="E2EFDA"),
        "fill_label":    PatternFill("solid", fgColor="F2F2F2"),
        "fill_none":     PatternFill(fill_type=None),
        "font_titulo":   Font(name="Calibri", bold=True, size=14, color="FFFFFF"),
        "font_subtit":   Font(name="Calibri", italic=True, size=10, color="FFFFFF"),
        "font_seccion":  Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
        "font_col_hdr":  Font(name="Calibri", bold=True, size=10, color="1F4E79"),
        "font_label":    Font(name="Calibri", bold=True, size=10, color="404040"),
        "font_valor":    Font(name="Calibri", size=10, color="000000"),
        "font_firma":    Font(name="Calibri", size=10, color="595959", italic=True),
        "alin_centro":   Alignment(horizontal="center", vertical="center", wrap_text=True),
        "alin_izq":      Alignment(horizontal="left", vertical="center", wrap_text=True),
        "borde":         borde_celda,
    }


def _construir_hoja_resumen(ws, df_ot: pd.DataFrame, df_inv: pd.DataFrame) -> None:
    """Construye la pestaña 'Resumen OTs' con encabezados y estilos corporativos."""
    E = _estilos_excel()
    ENCABEZADOS = [
        "ID Orden", "Código Equipo", "Tipo Mantenimiento",
        "Fecha", "Estado", "Repuestos Consumidos", "Cantidad", "Estado Final",
    ]
    ANCHOS = [12, 16, 22, 14, 16, 28, 12, 14]
    ultima_col = get_column_letter(len(ENCABEZADOS))

    ws.merge_cells(f"A1:{ultima_col}1")
    ws["A1"].value = "SISTEMA DE MANTENIMIENTO DE MAQUINARIA PESADA"
    ws["A1"].fill, ws["A1"].font, ws["A1"].alignment = E["fill_titulo"], E["font_titulo"], E["alin_centro"]
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{ultima_col}2")
    ws["A2"].value = f"Reporte de Órdenes de Trabajo  —  Generado: {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].fill, ws["A2"].font, ws["A2"].alignment = E["fill_seccion"], E["font_subtit"], E["alin_centro"]
    ws.row_dimensions[2].height = 16

    for col_i, (hdr, ancho) in enumerate(zip(ENCABEZADOS, ANCHOS), start=1):
        c = ws.cell(row=4, column=col_i, value=hdr)
        c.fill, c.font, c.alignment, c.border = E["fill_col_hdr"], E["font_col_hdr"], E["alin_centro"], E["borde"]
        ws.column_dimensions[get_column_letter(col_i)].width = ancho
    ws.row_dimensions[4].height = 20

    if df_ot.empty:
        ws.merge_cells(f"A5:{ultima_col}5")
        ws["A5"].value = "Sin órdenes de trabajo registradas."
        ws["A5"].alignment, ws["A5"].font = E["alin_centro"], E["font_firma"]
        return

    nombre_rep = df_inv.set_index("Codigo")["Nombre"].to_dict() if not df_inv.empty else {}

    for fila_i, (_, ot) in enumerate(df_ot.iterrows(), start=5):
        estado_final = "Completado" if ot["Estado"] == "Finalizado" else "En Proceso"
        fila_vals = [
            ot["ID_OT"], ot["Codigo_Equipo"], ot["Tipo_Mantenimiento"], str(ot["Fecha"]),
            ot["Estado"], nombre_rep.get(str(ot["Codigo_Repuesto"]), str(ot["Codigo_Repuesto"])),
            float(ot["Cantidad"]), estado_final,
        ]

        fill_fila = E["fill_fin"] if ot["Estado"] == "Finalizado" else (E["fill_alt"] if fila_i % 2 == 0 else None)

        for col_i, val in enumerate(fila_vals, start=1):
            c = ws.cell(row=fila_i, column=col_i, value=val)
            c.font, c.border = E["font_valor"], E["borde"]
            c.alignment = E["alin_centro"] if col_i in (1, 4, 5, 7, 8) else E["alin_izq"]
            if fill_fila:
                c.fill = fill_fila
        ws.row_dimensions[fila_i].height = 16


def _construir_ficha_tecnica(ws, df_ot: pd.DataFrame, df_eq: pd.DataFrame, df_inv: pd.DataFrame, id_ot_ficha: str | None) -> None:
    """Construye la pestaña 'Ficha Técnica' siguiendo la estructura formal de taller (Senati)."""
    E = _estilos_excel()

    for col, ancho in zip("ABCDEF", [20, 22, 20, 22, 12, 12]):
        ws.column_dimensions[col].width = ancho

    if id_ot_ficha is None or df_ot.empty:
        ws.merge_cells("A1:F1")
        ws["A1"].value = "No se seleccionó ninguna OT para la Ficha Técnica."
        ws["A1"].font, ws["A1"].alignment = E["font_firma"], E["alin_centro"]
        return

    fila_ot = df_ot[df_ot["ID_OT"] == id_ot_ficha]
    if fila_ot.empty:
        ws["A1"].value = f"OT '{id_ot_ficha}' no encontrada."
        return

    ot = fila_ot.iloc[0]
    eq = df_eq[df_eq["Codigo"].str.upper() == str(ot["Codigo_Equipo"]).upper()].iloc[0] if not df_eq.empty else None
    rep = df_inv[df_inv["Codigo"].str.upper() == str(ot["Codigo_Repuesto"]).upper()].